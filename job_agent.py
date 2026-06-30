import os
import re
import json
import asyncio
import logging
import datetime
import shutil
from typing import List, Optional
from pydantic import BaseModel, Field
import ollama
from openpyxl import Workbook, load_workbook

# Import json_repair to handle malformed 1.5B LLM outputs
import json_repair

from career_pages import get_career_url_map, get_career_config, get_all_navigation_instructions

# MCP Client SDK Imports
from mcp import ClientSession, StdioServerParameters
from mcp.client.stdio import stdio_client

# Setup logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger("job_agent")

# ==========================================
# LLM ROBUST PARSING UTILITY
# ==========================================
def parse_llm_response(raw_text: str) -> dict:
    """Safely extracts JSON from DeepSeek 1.5B conversational output."""
    try:
        # json_repair.loads acts just like json.loads but fixes broken JSON strings automatically
        parsed_data = json_repair.loads(raw_text)
        
        # Ensure it returns a dictionary, even if the model outputs a list of length 1
        if isinstance(parsed_data, list) and len(parsed_data) > 0:
            return parsed_data[0]
        elif isinstance(parsed_data, dict):
            return parsed_data
        return {}
    except Exception as e:
        logger.error(f"Failed to parse 1.5B model output: {e}")
        return {}

def calculate_job_score(llm_extracted_skills: list, user_profile_text: str) -> int:
    """Calculates match score strictly using Python to avoid 1.5B model hallucinations."""
    if not llm_extracted_skills or not user_profile_text:
        return 0
        
    user_text_lower = user_profile_text.lower()
    user_words = set(re.findall(r'\b\w+\b', user_text_lower))
    
    match_count = 0
    llm_skills_set = set(str(s).lower().strip() for s in llm_extracted_skills)
    
    for skill in llm_skills_set:
        # If it's a multi-word skill, just do a substring check
        if " " in skill and skill in user_text_lower:
            match_count += 1
        # If it's a single word, do a strict word boundary check
        elif skill in user_words:
            match_count += 1
            
    total_required = max(len(llm_skills_set), 1)
    
    # Calculate percentage out of 100
    score = int((match_count / total_required) * 100)
    return score

# ==========================================
# PYDANTIC SCHEMAS (Forced Structured Outputs w/ Defaults for 1.5B Resiliency)
# ==========================================
class JobDetails(BaseModel):
    is_job_posting: bool = Field(default=True, description="True if the text is a specific job posting/description, False if it is a general page, list, or error.")
    title: str = Field(default="N/A", description="The concise formal job title. MUST summarize and remove any extra description text (e.g., 'Software Engineer' instead of 'iconSoftware Engineer - We are seeking...'). Max 5 words.")
    company: str = Field(default="N/A", description="The name of the hiring organization or company, or 'N/A'")
    required_skills: List[str] = Field(default_factory=list, description="Core technical or domain skills listed. Empty list if none.")
    experience_level: str = Field(default="N/A", description="Required minimum years of experience or seniority level, or 'N/A'")
    explicit_salary: Optional[str] = Field(default=None, description="Explicit salary stated in text, e.g. '30 LPA', '₹20,00,000', '$120,000'")
    location: str = Field(default="N/A", description="Job location, e.g., 'Bengaluru', 'Remote', 'Mumbai', or 'N/A'")
    role: str = Field(default="N/A", description="Specific role category or specialization, e.g., 'Quant Developer', 'Data Scientist', 'Machine Learning Engineer', or 'N/A'")
    description: str = Field(default="N/A", description="Brief 2-3 sentence description of the job responsibilities and highlights.")
    posted_date: Optional[str] = Field(default="Unknown", description="The date the job was posted, if explicitly stated or inferred from text, e.g. '2026-05-10', '3 days ago', 'October 2025', or 'Unknown'.")

class SalaryEstimation(BaseModel):
    estimated_salary_range: str = Field(default="Unknown", description="Estimated salary range or amount")
    reason: str = Field(default="", description="Brief reason or source snippet for the estimation")

class JobMatchEvaluation(BaseModel):
    score: int = Field(default=0, description="Match score between 0 and 100 representing how well candidate profile matches the job requirements.")
    reason: str = Field(default="Evaluated by AI", description="A brief 1-2 sentence explanation of the rating.")
    is_curve_ball: bool = Field(default=False, description="True if the job is slightly outside the candidate's core domain/tech stack but represents a high-potential adjacent opportunity. False otherwise.")
    curve_ball_reason: Optional[str] = Field(default=None, description="If is_curve_ball is True, brief reason why this adjacent role is worth exploring.")

class RefinedConfig(BaseModel):
    search_queries: List[str] = Field(default_factory=list, description="List of 5-15 highly refined search query strings optimized for the candidate.")
    target_companies: List[str] = Field(default_factory=list, description="List of target companies.")

class JobRelevance(BaseModel):
    index: int = Field(default=0, description="The exact 'Index' number of the job from the input list.")
    is_relevant: bool = Field(default=True, description="True if the job title, company, or snippet suggests it is a match.")
    reason: str = Field(default="Matched target criteria", description="A short explanation of why it was kept or rejected.")

class BatchFilterResponse(BaseModel):
    evaluations: List[JobRelevance] = Field(default_factory=list, description="Evaluations for each job in the batch.")

# ==========================================
# CURRENCY & SALARY PARSER UTILITIES
# ==========================================
def parse_salary_to_target_currency(salary_str: str, target_currency: str = "USD") -> float:
    if not salary_str or salary_str.lower() in ["none", "n/a", "null", "unknown", "unspecified"]:
        return 0.0
    
    s = salary_str.lower()
    numbers = re.findall(r'\d+(?:,\d+)*(?:\.\d+)?', s)
    if not numbers:
        return 0.0
    
    val = float(numbers[0].replace(',', ''))
    if val == 0.0 and len(numbers) > 1:
        val = float(numbers[1].replace(',', ''))
    
    is_usd = any(c in s for c in ['$', 'usd', 'eur', 'gbp', '€', '£'])
    is_hourly = any(h in s for h in ['/hr', 'hour', 'hr'])
    is_monthly = any(m in s for m in ['/mo', 'month', 'pm'])
    
    if is_hourly:
        annual_val = val * 2000 
    elif is_monthly:
        annual_val = val * 12
    else:
        if re.search(r'\b\d+\s*k\b|\b\d+k\b', s) and val < 1000:
            annual_val = val * 1000
        else:
            annual_val = val
            
    has_lakh_indicator = any(re.search(rf'\b{word}\b', s) for word in ['lpa', 'lakh', 'lakhs', 'lac', 'lacs'])
    
    if has_lakh_indicator and val < 200:
        annual_val = annual_val * 100000
        
    target_currency = target_currency.upper()
    if target_currency == "USD" and not is_usd and annual_val > 500000:
        annual_val = annual_val / 83.0
    elif target_currency == "INR" and is_usd:
        annual_val = annual_val * 83.0
        
    return round(annual_val, 2)

def is_posted_over_6_months_ago(posted_date_str: str) -> bool:
    if not posted_date_str or posted_date_str.lower() in ["unknown", "n/a", "none", "null"]:
        return False
        
    s = posted_date_str.lower().strip()
    
    if "year" in s:
        return True
        
    month_match = re.search(r'(\d+)\s*month', s)
    if month_match:
        try:
            months = int(month_match.group(1))
            return months >= 6
        except Exception:
            pass
        
    if "week" in s or "day" in s or "hour" in s or "yesterday" in s or "today" in s:
        return False
        
    year_match = re.search(r'\b(20\d{2})\b', s)
    if year_match:
        try:
            year = int(year_match.group(1))
            today = datetime.date.today()
            if year < today.year - 1:
                return True
                
            parsed_date = None
            iso_match = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
            if iso_match:
                parsed_date = datetime.date(int(iso_match.group(1)), int(iso_match.group(2)), int(iso_match.group(3)))
            else:
                for fmt in ["%B %Y", "%b %Y", "%d %B %Y", "%d %b %Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"]:
                    try:
                        clean_s = re.sub(r'th|st|nd|rd', '', s)
                        parsed_dt = datetime.datetime.strptime(clean_s, fmt)
                        parsed_date = parsed_dt.date()
                        break
                    except ValueError:
                        continue
            
            if parsed_date:
                age_days = (today - parsed_date).days
                return age_days >= 180
        except Exception:
            pass
            
    return False

def update_scan_status(current_company: str, completed_companies: List[str], status: str = "running"):
    try:
        os.makedirs("data", exist_ok=True)
        status_path = os.path.join("data", "scan_status.json")
        with open(status_path, "w", encoding="utf-8") as f:
            json.dump({
                "status": status,
                "current_company": current_company,
                "completed_companies": completed_companies
            }, f, indent=2)
    except Exception as e:
        logger.error(f"Failed to update scan_status.json: {e}")

def clean_old_directories(data_dir: str = "data", max_days: int = 30):
    if not os.path.exists(data_dir):
        return
    today = datetime.date.today()
    logger.info(f"Running cleanup on '{data_dir}' folder. Deleting folders older than {max_days} days...")
    for item in os.listdir(data_dir):
        item_path = os.path.join(data_dir, item)
        if os.path.isdir(item_path):
            try:
                folder_date = datetime.datetime.strptime(item, "%Y-%m-%d").date()
                age = (today - folder_date).days
                if age > max_days:
                    logger.info(f"Deleting old folder: {item} (Age: {age} days)")
                    shutil.rmtree(item_path)
            except ValueError:
                pass

def save_to_excel_rolling(job: dict, excel_path: str):
    headers = [
        "Title", "Company", "URL", "Source", "Expected CTC", 
        "Required Skills", "Experience Level", "Match Score", "Match Reason", 
        "Location", "Role", "Description", "Posted Date"
    ]
    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = "Jobs"
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Jobs"
        ws.append(headers)
        
    ws.append([
        job.get("title", "N/A"),
        job.get("company", "N/A"),
        job.get("url", "N/A"),
        job.get("source", "N/A"),
        job.get("ctc", 0.0),
        ", ".join(job.get("required_skills", [])),
        job.get("experience_level", "N/A"),
        job.get("match_score", 0),
        job.get("match_reason", "N/A"),
        job.get("location", "N/A"),
        job.get("role", "N/A"),
        job.get("description", "N/A"),
        job.get("posted_date", "Unknown")
    ])
    try:
        wb.save(excel_path)
        logger.info(f"Incrementally saved job '{job.get('title')}' at '{job.get('company')}' to Excel.")
    except Exception as e:
        logger.error(f"Error saving rolling Excel: {e}")

async def update_preferences_profile(job_title: str, company: str, status: str, comment: str, preferences_path: str = ".resumes/preferences.md", model_name: str = "deepseek-r1:1.5b"):
    os.makedirs(os.path.dirname(preferences_path), exist_ok=True)
    current_profile = ""
    if os.path.exists(preferences_path):
        try:
            with open(preferences_path, "r", encoding="utf-8") as f:
                current_profile = f.read()
        except Exception as e:
            logger.error(f"Error reading preferences file: {e}")
            
    if not current_profile.strip():
        current_profile = "# Job Preferences Profile\n\n## Likes\n- None specified yet.\n\n## Dislikes\n- None specified yet.\n"

    prompt = f"""
    You are an AI assistant managing a candidate's career preferences.
    Update the candidate's current preferences markdown profile based on new user feedback.
    
    Current Preferences Profile:
    {current_profile}
    
    New Feedback:
    - Job: {job_title} at {company}
    - Rating: {status.upper()} (User gave a {"thumbs up" if status == "thumbs_up" else "thumbs down"})
    - User Comment on why: "{comment}"
    
    Please update the current profile. Integrate the new feedback.
    Return ONLY the complete updated Markdown text. Do not write code fences like ```markdown.
    """
    
    try:
        client = ollama.AsyncClient(host=load_config().get("ollama_host", "http://127.0.0.1:11434"))
        response = await asyncio.wait_for(
            client.chat(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are a professional assistant. Output ONLY the updated Markdown profile. Do not include chat explanations or markdown blocks."},
                    {"role": "user", "content": prompt}
                ],
                options={"num_ctx": 4096, "num_predict": 1024}
            ),
            timeout=120
        )
        updated_content = response['message']['content'].strip()
        
        if updated_content.startswith("```markdown"):
            updated_content = updated_content[11:]
        if updated_content.startswith("```"):
            updated_content = updated_content[3:]
        if updated_content.endswith("```"):
            updated_content = updated_content[:-3]
        updated_content = updated_content.strip()

        with open(preferences_path, "w", encoding="utf-8") as f:
            f.write(updated_content)
        logger.info(f"Preferences profile '{preferences_path}' successfully updated.")
    except Exception as e:
        logger.error(f"Error updating preferences profile: {e}")

def save_to_json_rolling(job: dict, json_path: str):
    jobs_list = []
    if os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                jobs_list = json.load(f)
        except Exception:
            pass
            
    exists = False
    for i, j in enumerate(jobs_list):
        if (j.get("id") and j.get("id") == job.get("id")) or (j.get("url") and j.get("url") == job.get("url")):
            jobs_list[i] = job
            exists = True
            break
            
    if not exists:
        jobs_list.append(job)
        
    try:
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(jobs_list, f, indent=2)
    except Exception as e:
        logger.error(f"Error saving rolling JSON: {e}")

def save_to_json_global(job: dict):
    os.makedirs("data", exist_ok=True)
    all_jobs_path = "data/all_jobs.json"
    jobs_list = []
    if os.path.exists(all_jobs_path):
        try:
            with open(all_jobs_path, "r", encoding="utf-8") as f:
                jobs_list = json.load(f)
        except Exception:
            pass
            
    exists = False
    for i, j in enumerate(jobs_list):
        if (j.get("id") and j.get("id") == job.get("id")) or (j.get("url") and j.get("url") == job.get("url")):
            jobs_list[i] = job
            exists = True
            break
            
    if not exists:
        jobs_list.append(job)
        
    try:
        with open(all_jobs_path, "w", encoding="utf-8") as f:
            json.dump(jobs_list, f, indent=2)
    except Exception as e:
        logger.error(f"Error saving global JSON: {e}")

def is_aggregator_list_page(url: str) -> bool:
    url_lower = url.lower()
    if any(p in url_lower for p in ["/jobs/view/", "/viewjob", "naukri.com/job-listings-", "greenhouse.io/", "lever.co/", "eightfold.ai/"]):
        return False
        
    patterns = [
        "/jobs/search", "search?", "query=", "?q=", "search_query=",
        "/jobs-in-", "/jobs-at-", "/jobs-for-",
        "indeed.com/q-", "indeed.com/l-",
        "simplyhired.co.in/search",
        "naukri.com/", "careerjet.co.in/",
        "amazon.jobs/content/", "amazon.jobs/en/search",
        "linkedin.com/jobs/", "glassdoor.co.in/job/"
    ]
    
    for pattern in patterns:
        if pattern in url_lower:
            return True
            
    return False

def is_valid_job_post(url: str) -> bool:
    """Strictly checks if a URL is likely an actual job post, not a blog/article/general page."""
    u = url.lower()
    
    # 1. Known ATS Platforms
    ats_domains = [
        "greenhouse.io", "lever.co", "workdayjobs.com", "myworkdayjobs.com",
        "ashbyhq.com", "icims.com", "smartrecruiters.com", "breezy.hr", "bamboohr.com", "eightfold.ai"
    ]
    if any(ats in u for ats in ats_domains):
        return True
        
    # 2. Known job board paths
    job_board_paths = [
        "linkedin.com/jobs/view/", "indeed.com/viewjob", "naukri.com/job-listings-"
    ]
    if any(p in u for p in job_board_paths):
        return True
        
    # 3. Generic career portals but with job-specific paths
    # Matches /job/123, /jobs/123, /careers/job/123, /role/123, /position/123
    generic_job_regex = r"/(?:jobs?|careers?|roles?|positions?|opportunities|openings?)/[^/]*\d+[^/]*$"
    if re.search(generic_job_regex, u):
        return True
        
    # 4. Sometimes they have alphanumeric GUIDs
    guid_job_regex = r"/(?:jobs?|careers?)/[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}"
    if re.search(guid_job_regex, u):
        return True
        
    return False
    
    for pattern in patterns:
        if pattern in url_lower:
            return True
            
    return False

async def process_discovered_url(url: str, title: str, snippet: str, session, config: dict, model_name: str, salary_threshold: float, resume_text: str, preferences_text: str, daily_dir: str, company_hint: str = "") -> Optional[dict]:
    jobs_json_path = os.path.join(daily_dir, "jobs.json")
    if os.path.exists(jobs_json_path):
        try:
            with open(jobs_json_path, "r", encoding="utf-8") as f:
                existing_jobs = json.load(f)
                if any(j.get("url") == url and "match_reason" in j for j in existing_jobs):
                    logger.info(f"Already deeply processed URL {url}. Skipping.")
                    return None
        except Exception:
            pass

    source_domain = "Web Search"
    domain_match = re.search(r'https?://(?:www\.)?([^/]+)', url)
    if domain_match:
        source_domain = domain_match.group(1)
        
    cleaned_text = ""
    try:
        page_res = await session.call_tool("fetch_web_page", arguments={"url": url})
        cleaned_text = page_res.content[0].text
    except Exception as e:
        logger.warning(f"Failed to scrape webpage directly: {e}")
        
    if not cleaned_text.strip() or cleaned_text.startswith("Error fetching") or len(cleaned_text.strip()) < 100:
        if snippet:
            logger.warning("Falling back to search snippet...")
            cleaned_text = f"Title: {title}\nJob Posting Snippet Details:\n{snippet}"
        else:
            logger.warning(f"Skipping {url} due to fetch error.")
            return None

    nav_context = ""
    if company_hint:
        cfg = get_career_config(company_hint)
        if cfg:
            nav_context = f"\n\nCAREER PORTAL NAVIGATION CONTEXT for {cfg.company}:\n{cfg.model_instructions.strip()}\n"

    # Allow sufficient context window for 1.5B models
    MAX_JOB_CHARS = 12000
    truncated_job_text = cleaned_text[:MAX_JOB_CHARS]

    # Simplified extraction prompt
    extract_prompt = f"Extract only these fields as JSON: is_job_posting,title,company,required_skills,experience_level,explicit_salary,location,role,description,posted_date. If it is NOT a job posting, set is_job_posting to false.\nCRITICAL: Ensure 'title' is a CONCISE 2-5 word job title. Remove any prefixed HTML garbage like 'icon' or appended descriptions.\n\nText:\n{truncated_job_text}"
    
    try:
        client = ollama.AsyncClient(host=config.get("ollama_host", "http://127.0.0.1:11434"))
        response = await asyncio.wait_for(
            client.chat(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are an extraction bot. Output ONLY valid JSON matching the requested keys. No extra text."},
                    {"role": "user", "content": extract_prompt}
                ],
                format=JobDetails.model_json_schema(),
                options={"num_ctx": 4096, "num_predict": 1024}
            ),
            timeout=120
        )
        content = response.get('message', {}).get('content', '')
        
        try:
            # 1.5B Model Resiliency: Use json_repair
            repaired_dict = parse_llm_response(content)
            
            # Fallback to company_hint if model missed it
            if repaired_dict.get("company") in [None, "N/A", "", "Unknown"] and company_hint:
                repaired_dict["company"] = company_hint
                
            job_metrics = JobDetails(**repaired_dict)
            
            # Clean up trailing JSON artifacts in title
            if job_metrics.title:
                job_metrics.title = re.sub(r'\},?\s*$', '', job_metrics.title).strip()
                
        except Exception as ve:
            logger.error(f"Failed to load into Pydantic model: {ve}")
            return None

    except Exception as e:
        err_str = str(e).lower()
        if "connection" in err_str or "urllib3" in err_str or "http" in err_str or "refused" in err_str or "timeout" in err_str:
            logger.critical(f"Ollama server is unreachable or timed out during extraction! Skipping url: {url} - {e}")
        else:
            logger.error(f"Error calling Ollama extraction: {e}")
        return None

    if not job_metrics.is_job_posting:
        logger.info("Webpage is not a job posting. Skipping.")
        return None

    if job_metrics.posted_date and is_posted_over_6_months_ago(job_metrics.posted_date):
        logger.info(f"⛔ Gated: Job posted date '{job_metrics.posted_date}' is 6 months or more ago. Skipping.")
        return None
        
    logger.info(f"Found Job: '{job_metrics.title}' at '{job_metrics.company}'")
    
    location = job_metrics.location if job_metrics.location else "N/A"
    role = job_metrics.role if job_metrics.role else "N/A"
    description = job_metrics.description if job_metrics.description else "N/A"

    estimated_pay = 0.0
    if job_metrics.explicit_salary:
        estimated_pay = parse_salary_to_target_currency(job_metrics.explicit_salary, config.get('currency', 'USD'))
        logger.info(f"Parsed Explicit Salary: {job_metrics.explicit_salary} -> ~{estimated_pay} {config.get('currency', 'USD')}")

    if not resume_text:
        logger.info("Skipping resume match evaluation because no resumes were loaded.")
        return None
        
    logger.info("Evaluating candidate alignment against resumes and preference profile...")
    career_nav_guide = get_all_navigation_instructions()
    resume_snip_len = config.get('resume_max_chars_in_prompt', 1000) if isinstance(config, dict) else 1000
    trimmed_resume = (resume_text[:resume_snip_len] + '...') if resume_text and len(resume_text) > resume_snip_len else (resume_text or '')

    match_prompt = f"""
    Evaluate structural alignment between Candidate Resumes, Preferences, and the Target Job details.
    
    Candidate Resumes:
    {trimmed_resume}
    
    Candidate Preferences & Dislikes:
    {preferences_text if preferences_text else "No specific preferences set yet."}
    
    Target Job:
    - Title: {job_metrics.title}
    - Company: {job_metrics.company}
    - Skills: {', '.join(job_metrics.required_skills)}
    - Experience/Seniority: {job_metrics.experience_level}
    - Location: {location}
    - Role: {role}
    
    INSTRUCTIONS:
    1. Output JSON. Generate a 'score' from 0-100 indicating how well the candidate's skills match the job requirements. Generate 'reason', 'is_curve_ball', and 'curve_ball_reason'.
    2. If the job is NOT in {config.get('target_location', 'Remote')} or Remote, state this in the reason.
    3. If the job violates candidate dislikes, state this in the reason.
    4. If it's a high-potential adjacent opportunity, set is_curve_ball to True and explain.
    """
    
    try:
        client = ollama.AsyncClient(host=config.get("ollama_host", "http://127.0.0.1:11434"))
        match_res = await asyncio.wait_for(
            client.chat(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are a professional recruiter. Output JSON matching the schema."},
                    {"role": "user", "content": match_prompt}
                ],
                format=JobMatchEvaluation.model_json_schema(),
                options={"num_ctx": 4096, "num_predict": 256}
            ),
            timeout=120
        )
        # 1.5B Model Resiliency
        match_dict = parse_llm_response(match_res['message']['content'])
        
        # Rely on the semantic score generated by the 1.5B LLM
        # to arrange jobs accurately based on relevancy, instead of rigid Python checks.
        pass
        
        match_data = JobMatchEvaluation(**match_dict)
        logger.info(f"Candidate Match Score (Python Calculated): {match_data.score}% | Curve Ball: {match_data.is_curve_ball}")
    except Exception as e:
        err_str = str(e).lower()
        if "connection" in err_str or "urllib3" in err_str or "http" in err_str or "refused" in err_str or "timeout" in err_str:
            logger.critical(f"Ollama server is unreachable or timed out during match evaluation! Skipping url: {url} - {e}")
        else:
            logger.error(f"Error evaluating match: {e}")
        return None

    job_id = re.sub(r'[^a-z0-9]', '_', f"{job_metrics.company}_{job_metrics.title}".lower())
    
    job_result = {
        "id": job_id,
        "title": job_metrics.title,
        "company": job_metrics.company,
        "url": url,
        "source": source_domain,
        "ctc": estimated_pay,
        "required_skills": job_metrics.required_skills,
        "experience_level": job_metrics.experience_level,
        "location": location,
        "role": role,
        "description": description,
        "posted_date": job_metrics.posted_date if job_metrics.posted_date else "Unknown",
        "match_score": match_data.score,
        "match_reason": match_data.reason,
        "is_curve_ball": match_data.is_curve_ball,
        "curve_ball_reason": match_data.curve_ball_reason,
        "status": "unrated",
        "feedback_comment": ""
    }
    
    save_to_json_rolling(job_result, jobs_json_path)
    save_to_json_global(job_result)
    excel_path = os.path.join(daily_dir, "jobs.xlsx")
    save_to_excel_rolling(job_result, excel_path)
    
    return job_result

# ==========================================
# MAIN EXECUTION ROUTINE
# ==========================================
DEFAULT_CONFIG = {
    "ollama_model": "deepseek-r1:1.5b",
    "target_compensation_threshold": 100000,
    "currency": "USD",
    "target_location": "Remote",
    "disliked_companies": ["infosys", "wipro", "tcs", "cognizant"],
    "resumes_dir": ".resumes",
    "smart_filter_batch_size": 1, 
    "search_queries": [
        "data scientist",
        "applied scientist",
        "quant analyst",
        "site:linkedin.com/jobs/view \"quant researcher\"",
        "site:linkedin.com/jobs/view \"data scientist\""
    ],
    "headless": True,
    "user_data_dir": ".browser_profile",
    "preferences_path": ".resumes/preferences.md",
    "target_companies": [
        "Tower Research Capital",
        "Jane Street",
        "D. E. Shaw",
        "Google",
        "Uber"
    ],
    "direct_job_urls": [],
    "jobs_digest_path": "jobs_digest.md",
    "restrict_to_target_companies": False,
    "max_additional_companies": 30,
    "max_jobs_to_process": 50,
    "company_career_pages": {},
    "scan_interval_hours": 6,
    "career_scrape_max_per_company": 15
}

def load_config() -> dict:
    config_path = "config.json"
    config = {}
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                content = f.read().strip()
                if content:
                    config = json.loads(content)
        except Exception as e:
            logger.error(f"Error reading or parsing config.json: {e}")
            config = {}
            
    modified = False
    for k, v in DEFAULT_CONFIG.items():
        if k not in config:
            config[k] = v
            modified = True
            
    try:
        hardcoded = get_career_url_map()
        if hardcoded:
            career_pages = config.get("company_career_pages", {}) or {}
            target_companies = config.get("target_companies", []) or []
            for co, url in hardcoded.items():
                if co not in career_pages:
                    career_pages[co] = url
                    modified = True
            config["company_career_pages"] = career_pages
    except Exception as e:
        logger.error(f"Failed merging hardcoded career urls into config: {e}")
            
    if modified:
        try:
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=2)
            logger.info("config.json was missing or empty keys. Restored defaults successfully.")
        except Exception as e:
            logger.error(f"Failed to save default config.json: {e}")
            
    return config

async def optimize_search_queries_and_config(resume_text: str, preferences_text: str, config: dict, model_name: str) -> dict:
    logger.info("Starting proactive search query and target company optimization...")
    if not resume_text and not preferences_text:
        logger.warning("No resume or preferences available for search optimization. Skipping.")
        return config

    prompt = f"""
    You are an expert recruiter. Analyze the candidate's Resume and Preferences Profile.
    Output JSON containing optimized "search_queries" and "target_companies".

    Candidate Resume Profile:
    {resume_text[:2000]}

    Candidate Preferences Profile (Likes & Dislikes):
    {preferences_text}

    Current Target Companies:
    {json.dumps(config.get("target_companies", [])[:10], indent=2)}
    """

    try:
        client = ollama.AsyncClient(host=config.get("ollama_host", "http://127.0.0.1:11434"))
        response = await asyncio.wait_for(
            client.chat(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are a professional recruiting configurator. Return JSON matching the schema precisely. Validate location and dislikes."},
                    {"role": "user", "content": prompt}
                ],
                format=RefinedConfig.model_json_schema(),
                options={"num_ctx": 4096, "num_predict": 512}
            ),
            timeout=120
        )
        
        dict_val = parse_llm_response(response['message']['content'])
        refined_data = RefinedConfig(**dict_val)
        
        cleaned_queries = []
        for q in refined_data.search_queries:
            if not isinstance(q, str): continue
            q_clean = q.strip()
            if len(q_clean) > 3 and not any(char in q_clean for char in [';', '|', '&', '$', '`', '<', '>']):
                cleaned_queries.append(q_clean)
        
        cleaned_companies = []
        for c in refined_data.target_companies:
            if not isinstance(c, str): continue
            c_clean = c.strip()
            if len(c_clean) > 2 and not any(char in c_clean for char in [';', '|', '&', '$', '`', '<', '>']):
                c_lower = c_clean.lower()
                disliked = False
                for term in config.get("disliked_companies", ["infosys", "wipro", "tcs", "cognizant"]):
                    if term in c_lower:
                        disliked = True
                        break
                if not disliked:
                    cleaned_companies.append(c_clean)
                    
        if cleaned_queries:
            config["search_queries"] = cleaned_queries[:25]
        if cleaned_companies:
            existing_companies = config.get("target_companies", [])
            disliked_terms = config.get("disliked_companies", ["infosys", "wipro", "tcs", "cognizant"])
            filtered_existing = [c for c in existing_companies if not any(t in c.lower() for t in disliked_terms)]
            merged_companies = list(filtered_existing)
            for c in cleaned_companies:
                if c.lower() not in [x.lower() for x in merged_companies]:
                    merged_companies.append(c)
            config["target_companies"] = merged_companies[:30]
            
        try:
            # Removed permanent overwrite to prevent LLM hallucinations from destroying user config
            # with open("config.json", "w", encoding="utf-8") as f:
            #     json.dump(config, f, indent=2)
            pass
        except Exception as write_err:
            logger.error(f"Failed to save optimized config.json: {write_err}")
            
    except Exception as e:
        logger.error(f"Failed to optimize search configuration: {e}")
        
    return config

async def filter_jobs_batch_smart(job_list: List[dict], resume_text: str, preferences_text: str, model_name: str) -> List[dict]:
    if not job_list: return []
    logger.info(f"Starting smart batch filtering of {len(job_list)} compiled job listings...")
    if not resume_text and not preferences_text: return job_list

    filtered_jobs = []
    config_local = load_config()
    # Enforce Batch Size of 1 to prevent Context Bleeding in 1.5B model
    batch_size = config_local.get("smart_filter_batch_size", 1) 
    resume_snip_len = config_local.get("resume_max_chars_in_prompt", 1000)
    
    for i in range(0, len(job_list), batch_size):
        batch = job_list[i:i + batch_size]
        
        jobs_input_list = []
        for idx, job in enumerate(batch, 1):
            jobs_input_list.append(
                f"Job Index: {idx}\n"
                f"Title: {job.get('title', 'N/A')}\n"
                f"Company: {job.get('company_hint', 'Unknown')}\n"
                f"Snippet: {job.get('snippet', '')[:300]}\n"
                f"Source: {job.get('source', 'Web Search')}\n"
                f"---"
            )
        jobs_input_str = "\n".join(jobs_input_list)
        trimmed_resume = resume_text[:resume_snip_len] if resume_text else ""
            
        prompt = f"""
        Candidate Summary: {trimmed_resume}
        Preferences: {preferences_text}

        Jobs:
        {jobs_input_str}

        INSTRUCTIONS:
        Output JSON matching the schema indicating if the job matches the candidate's target profile ({config_local.get('target_location', 'Remote')}/Remote locations, non-disliked companies).
        """
        
        try:
            client = ollama.AsyncClient(host=config_local.get("ollama_host", "http://127.0.0.1:11434"))
            response = await asyncio.wait_for(
                client.chat(
                    model=model_name,
                    messages=[
                        {"role": "system", "content": "You are a precise screening agent. Filter out irrelevant roles and locations. Output JSON."},
                        {"role": "user", "content": prompt}
                    ],
                    format=BatchFilterResponse.model_json_schema(),
                    options={"num_ctx": 4096, "num_predict": 512}
                ),
                timeout=120
            )
            dict_val = parse_llm_response(response['message']['content'])
            if isinstance(dict_val, list):
                dict_val = {"evaluations": dict_val}
            result = BatchFilterResponse(**dict_val)
            
            relevance_map = {eval_item.index: eval_item for eval_item in result.evaluations}
            
            for idx, job in enumerate(batch, 1):
                eval_item = relevance_map.get(idx)
                if eval_item:
                    if eval_item.is_relevant:
                        logger.info(f"✅ Kept: '{job.get('title')}' - {eval_item.reason}")
                        filtered_jobs.append(job)
                    else:
                        logger.info(f"❌ Skipped: '{job.get('title')}' - {eval_item.reason}")
                else:
                    filtered_jobs.append(job)
                    
        except Exception as batch_err:
            logger.error(f"Error filtering batch of jobs: {batch_err}. Defaulting to keep.")
            filtered_jobs.extend(batch)

    return filtered_jobs

async def verify_open_jobs():
    """Reads data/all_jobs.json and probes URLs to drop closed jobs."""
    all_jobs_path = "data/all_jobs.json"
    if not os.path.exists(all_jobs_path):
        return
        
    try:
        with open(all_jobs_path, "r", encoding="utf-8") as f:
            all_jobs = json.load(f)
    except Exception as e:
        logger.error(f"Could not load all_jobs.json: {e}")
        return
        
    logger.info(f"Verifying {len(all_jobs)} historical jobs to see if they are still open...")
    
    def check_url_sync(job):
        url = job.get("url")
        if not url: return None
        try:
            headers = {"User-Agent": "Mozilla/5.0"}
            resp = requests.get(url, headers=headers, timeout=10, allow_redirects=True)
            if resp.status_code in (404, 410):
                logger.info(f"Job {url} returned {resp.status_code}, marking closed.")
                return None
            elif resp.status_code >= 400:
                logger.warning(f"Job {url} returned {resp.status_code}, conservatively keeping it (likely bot protection).")
                return job
                
            text_lower = resp.text.lower()
            closed_phrases = [
                "no longer accepting applications",
                "this position has been filled",
                "job is closed",
                "job not found",
                "role is no longer available"
            ]
            if any(p in text_lower for p in closed_phrases):
                logger.info(f"Job {url} contains closed phrase, marking closed.")
                return None
            
            final_url = resp.url
            if "greenhouse.io" in final_url and "/jobs/" not in final_url:
                return None
            if "lever.co" in final_url and "/postings/" not in final_url and url != final_url:
                return None
                
            return job
        except Exception:
            return job

    async def check_url(job):
        return await asyncio.to_thread(check_url_sync, job)

    tasks = [check_url(j) for j in all_jobs]
    results = await asyncio.gather(*tasks)
    open_jobs = [r for r in results if r]
    
    try:
        with open(all_jobs_path, "w", encoding="utf-8") as f:
            json.dump(open_jobs, f, indent=2)
        logger.info(f"Verification complete. Kept {len(open_jobs)} / {len(all_jobs)} jobs.")
    except Exception as e:
        logger.error(f"Could not save verified all_jobs.json: {e}")

async def run_agent():
    """Main orchestration loop: Read resume, optimize config, fetch & filter jobs."""
    import time
    logger.info("Initializing Job Hunter Agent...")
    
    await verify_open_jobs()
    
    config = load_config()
    model_name = config.get("ollama_model", "deepseek-r1:1.5b")
    salary_threshold = config.get("target_compensation_threshold", 0)
    resumes_dir = config.get("resumes_dir", ".resumes")
    search_queries = config.get("search_queries", [])
    preferences_path = config.get("preferences_path", ".resumes/preferences.md")
    
    clean_old_directories(data_dir="data", max_days=30)
    
    today_str = datetime.date.today().isoformat()
    daily_dir = os.path.join("data", today_str)
    os.makedirs(daily_dir, exist_ok=True)
    jobs_json_path = os.path.join(daily_dir, "jobs.json")

    markdown_contents = []
    if os.path.exists(resumes_dir):
        for filename in os.listdir(resumes_dir):
            if filename.endswith(".md"):
                file_path = os.path.join(resumes_dir, filename)
                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        markdown_contents.append(f.read())
                except Exception as e:
                    pass
    preferences_text = "\n".join(markdown_contents)

    python_exe = os.path.join(".venv", "Scripts", "python.exe")
    if not os.path.exists(python_exe):
        python_exe = "python"
        
    server_params = StdioServerParameters(command=python_exe, args=["mcp_server.py"])
    
    async with stdio_client(server_params) as (read, write):
        async with ClientSession(read, write) as session:
            await session.initialize()
            resumes_result = await session.call_tool("read_resumes", arguments={"directory": resumes_dir})
            resume_text = resumes_result.content[0].text
            if "No readable resume files" in resume_text: resume_text = ""

            config = await optimize_search_queries_and_config(resume_text, preferences_text, config, model_name)
            search_queries = config.get("search_queries", [])
            target_companies = config.get("target_companies", [])

            direct_urls = config.get("direct_job_urls", [])
            job_list = [{"url": u, "title": "Direct Link", "snippet": "", "priority": 1} for u in direct_urls]

            completed_companies = []
            career_pages = config.get("company_career_pages", {})
            restrict_to_target = config.get("restrict_to_target_companies", False)
            max_additional = config.get("max_additional_companies", 30)
            max_to_process = config.get("max_jobs_to_process", 25)
            career_max = config.get("career_scrape_max_per_company", 15)
            
            base_roles = ["data scientist", "quant developer", "software engineer", "analyst", "quant", "researcher", "applied", "scientist"]

            async def filter_company_jobs(company_name, company_jobs_list):
                if not company_jobs_list: return []
                try:
                    filtered = await filter_jobs_batch_smart(company_jobs_list, resume_text, preferences_text, model_name)
                    for fj in filtered:
                        if 'company' not in fj or not fj.get('company'): fj['company'] = fj.get('company_hint', 'N/A')
                        if 'match_score' not in fj: fj['match_score'] = 0
                        if 'ctc' not in fj: fj['ctc'] = 0.0
                        if 'is_curve_ball' not in fj: fj['is_curve_ball'] = False
                    return filtered
                except Exception as e:
                    logger.error(f"Async filter for company {company_name} failed: {e}")
                    return []
            
            filter_tasks = []

            logger.info(f"Phase 0: Scraping targeted career portals for {target_companies}")
            role_kw_str = ", ".join(base_roles[:8])
            scraped_companies = set()
            for company in target_companies:
                update_scan_status(current_company=company, completed_companies=completed_companies)
                company_jobs = []
                try:
                    scrape_res = await session.call_tool("scrape_company_career_page", arguments={"company": company, "role_keywords": role_kw_str, "target_location": config.get("target_location", ""), "max_jobs": career_max})
                    scrape_text = scrape_res.content[0].text
                    if scrape_text.startswith("No hardcoded"): continue
                    scraped_companies.add(company.lower())
                    for block in re.split(r"Job \d+:", scrape_text)[1:]:
                        if "URL:" not in block: continue
                        url_m = re.search(r"URL:\s*(https?://\S+)", block)
                        title_m = re.search(r"Title:\s*(.*)", block)
                        if url_m:
                            url = url_m.group(1).strip()
                            if not any(j["url"] == url for j in job_list) and not any(j["url"] == url for j in company_jobs):
                                company_jobs.append({"url": url, "title": title_m.group(1).strip() if title_m else f"{company} Job", "snippet": "", "priority": 0, "company_hint": company})
                except Exception as e:
                    logger.error(f"Career portal scrape failed for '{company}': {e}")
                finally:
                    completed_companies.append(company)
                    update_scan_status(current_company="", completed_companies=completed_companies)
                    filter_tasks.append(asyncio.create_task(filter_company_jobs(company, company_jobs)))

            final_initial_jobs = list(job_list) # Includes direct URLs
            if filter_tasks:
                results_list = await asyncio.gather(*filter_tasks)
                for res in results_list:
                    final_initial_jobs.extend(res)
            
            final_initial_jobs.sort(key=lambda x: x.get("priority", 99))
            job_list = final_initial_jobs
            
            # Write the stub jobs so UI updates immediately
            with open(jobs_json_path, 'w', encoding='utf-8') as jf:
                json.dump(job_list, jf, indent=2)
            
            processed_count = 0
            daily_jobs = []

            # Process jobs concurrently
            semaphore = asyncio.Semaphore(3)

            async def sem_process_url(idx, job):
                async with semaphore:
                    try:
                        return await process_discovered_url(
                            url=job["url"], title=job["title"], snippet=job["snippet"],
                            session=session, config=config, model_name=model_name,
                            salary_threshold=salary_threshold, resume_text=resume_text,
                            preferences_text=preferences_text, daily_dir=daily_dir, company_hint=job.get("company_hint", "")
                        )
                    except Exception as e:
                        logger.error(f"Error processing {job['url']}: {e}")
                        return None

            tasks = []
            for idx, job in enumerate(job_list[:max_to_process], 1):
                tasks.append(sem_process_url(idx, job))

            results = await asyncio.gather(*tasks)
            
            for res in results:
                if res:
                    processed_count += 1
                    daily_jobs.append(res)
            
            if os.path.exists(jobs_json_path):
                with open(jobs_json_path, "r", encoding="utf-8") as f: final_jobs = json.load(f)
                final_jobs.sort(key=lambda x: x.get("match_score", 0), reverse=True)
                with open(jobs_json_path, "w", encoding="utf-8") as f: json.dump(final_jobs, f, indent=2)

            update_scan_status(current_company="", completed_companies=completed_companies, status="completed")

if __name__ == "__main__":
    asyncio.run(run_agent())